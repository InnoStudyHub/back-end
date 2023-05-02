import urllib.request
from io import BytesIO

import re
import PIL.Image
from PIL.ImageFile import ImageFile
from django.core.files.uploadedfile import UploadedFile
from openpyxl.reader.excel import load_workbook
from openpyxl_image_loader import SheetImageLoader
from rest_framework.exceptions import ValidationError

from deck.dtos.image_dto import Image
from deck.models import Deck
from studyhub.settings import logger, s3_client, AWS_STORAGE_BUCKET_NAME
from django.core.files.storage import default_storage


def isImage(image):
    return image.content_type is not None and image.content_type.split('/')[0] == 'image'


def uploadPublicFileToStorage(bucket_name, contents, destination_blob_name):
    logger.info("Uploading image to storage")
    s3_client.put_object(
        Bucket=bucket_name,
        Key=destination_blob_name,
        Body=contents,
        ContentType='image/jpg',
        ACL='public-read',
    )
    return f'https://storage.yandexcloud.net/{bucket_name}/{destination_blob_name}'


def uploadImage(files, deck_id, image_key):
    file = files.get(image_key)
    deck = Deck.objects.get(deck_id=deck_id)
    folder_name = deck.folder.folder_name
    extension = file.name.split('.')[-1]
    store_path = f'decks/{folder_name}/{deck.semester}/{deck_id}_{deck.deck_name}/{image_key}.{extension}'
    return uploadPublicFileToStorage(AWS_STORAGE_BUCKET_NAME, file.read(), store_path)


def getQuestionImageUrl(question_image_key, files, deck_id):
    if question_image_key:
        return uploadImage(files=files, deck_id=deck_id, image_key=question_image_key)
    return None


def getAnswerImageUrls(files, answer_image_keys, deck_id):
    answer_image_urls = []
    for answer_key in answer_image_keys:
        answer_image_urls.append(uploadImage(files=files,
                                             deck_id=deck_id,
                                             image_key=answer_key))
    return answer_image_urls


def parseGoogleSheet(url_old):
    logger.info(f"Parsing google sheet: {url_old}")
    match_key = re.search(r'docs.google.com/spreadsheets/d/[^/]+', url_old)
    match_gid = re.search(r'gid=[0-9]+', url_old)
    if not match_key:
        raise ValidationError("Not correct url")
    sheet_key = match_key.group().split('/')[-1]
    url = f'https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx'
    if match_gid:
        url += f'&{str(match_gid.group())}'
    logger.info(f"Parsed url {url}")
    file = urllib.request.urlopen(url).read()
    doc = load_workbook(filename=BytesIO(file))
    sheet = doc[doc.sheetnames[0]]

    image_loader = SheetImageLoader(sheet)

    cards = []
    for row in range(1, sheet.max_row + 1):
        if row == 1:
            continue

        question_cell = sheet.cell(row, 1)
        question_image_cell = sheet.cell(row, 2)
        answer_cell = sheet.cell(row, 3)

        answer_text = 'No answer'

        if not isinstance(question_cell.value, str):
            continue

        if question_image_cell.value is not None:
            raise ValidationError(f"In row-{row} question_image is not exist or not image")

        if answer_cell.value is not None and isinstance(answer_cell.value, str):
            answer_text = answer_cell.value

        answer_images = []
        for col in range(4, sheet.max_column + 1):
            answer_image_cell = sheet.cell(row, col)
            if answer_image_cell.value is not None:
                raise ValidationError(f"In row-{row} and column-{col} answer image is not image")
            if image_loader.image_in(answer_image_cell.coordinate):
                logger.info(f"image cell-{answer_image_cell.coordinate}")
                answer_images.append(image_loader.get(answer_image_cell.coordinate))

        question_image = None
        if image_loader.image_in(question_image_cell.coordinate):
            question_image = image_loader.get(question_image_cell.coordinate)

        cards.append({"question_text": question_cell.value, "question_image": question_image,
                      "answer_text": answer_text, "answer_images": answer_images})
    logger.info(f"Found {len(cards)} cards")
    return cards


def getCardDataAndFiles(card, key_count):
    card_data = {}
    files = {}
    if card.get('question_text'):
        card_data['question_text'] = card['question_text']
    if card.get('answer_text'):
        card_data['answer_text'] = card['answer_text']

    count = 0
    if card.get('question_image'):
        count += 1
        question_image_key = f"image_{key_count}_{count}"
        card_data['question_image_key'] = question_image_key
        files[question_image_key] = card['question_image']

    answer_image_keys = []
    for answer_image in card.get('answer_images'):
        count += 1
        answer_image_key = f"image_{key_count}_{count}"
        answer_image_keys.append(answer_image_key)
        files[answer_image_key] = answer_image

    card_data['answer_image_keys'] = answer_image_keys

    return {'card_data': card_data, 'files': files}


def toImage(file, key):
    if isinstance(file, bytes):
        b = BytesIO(file)
        b.seek(0)
        img = PIL.Image.new(mode="RGB", size=(256, 256))
        img.save(b, "jpeg")
        img.show()
        return Image(key + ".jpg", file, "image/jpg")
    elif isinstance(file, UploadedFile):
        if file.content_type == 'application/json':
            return Image(key + ".jpg", file.read(), "image/jpg")
        return Image(file.name, file.read(), file.content_type)
    elif isinstance(file, ImageFile):
        b = BytesIO()
        file.save(b, file.format)
        file.close()
        return Image(key + f".{file.format}", b.getvalue(), "image/jpeg")
