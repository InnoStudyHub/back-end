import logging
import urllib.request
from io import BytesIO

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

from rest_framework.exceptions import ValidationError

from deck.models import Deck
from deck.serializers.deck_serializer import DeckDetailSerializer, DeckPreviewSerializer

logger = logging.getLogger(__name__)

def getDeckData(deck, user):
    logger.info("Try to create DeckData")
    cards = deck.card_set.all()
    data = deck.__dict__
    data['cards'] = []
    data['is_favourite'] = user.favourite_decks.contains(deck)
    for card in cards:
        data['cards'].append(card.__dict__)
    serializer = DeckDetailSerializer(data=data)
    if not serializer.is_valid(raise_exception=True):
        raise ValidationError(serializer.error_messages)
    logger.info(f"DeckData successfully created: {serializer.data}")
    return serializer.data


def getDeckPreview(deck, user):
    logger.info("Try to create DeckDataPreview")
    cards = deck.card_set.all()
    data = deck.__dict__
    data['cards'] = len(cards)
    data['is_favourite'] = user.favourite_decks.contains(deck)
    serializer = DeckPreviewSerializer(data=data)
    if not serializer.is_valid(raise_exception=True):
        raise ValidationError(serializer.error_messages)
    logger.info(f"DeckDataPreview successfully created: {serializer.data}")
    return serializer.data


def getDecks(filter):
    return Deck.objects.all().filter(**filter)

def parseGoogleSheet(url):
    file = urllib.request.urlopen(url).read()
    doc = load_workbook(filename=BytesIO(file))
    sheet = doc[doc.sheetnames[0]]
    image_loader = SheetImageLoader(sheet)
    imageCells = list(image_loader.images.keys())

    image = image_loader.get('B13')
